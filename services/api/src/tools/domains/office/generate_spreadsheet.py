"""
Generate Spreadsheet Tool - Generación de archivos Excel

Crea archivos Excel (.xlsx) a partir de datos tabulares/JSON.
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List
import structlog

logger = structlog.get_logger()

# Workspace path for storing spreadsheets
if os.path.exists("/workspace") and os.access("/workspace", os.W_OK):
    WORKSPACE_PATH = Path("/workspace/spreadsheets")
else:
    # Fallback para entorno local de desarrollo
    WORKSPACE_PATH = Path(os.path.expanduser("~/workspace/spreadsheets"))


async def generate_spreadsheet(
    data: List[Dict[str, Any]],
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    file_name: Optional[str] = None,
    conversation_id: Optional[str] = None,
    agent_id: Optional[str] = "data_agent",
    metadata: Optional[Dict[str, Any]] = None,
    _user_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Genera un archivo Excel (.xlsx) a partir de datos tabulares.
    
    Args:
        data: Lista de diccionarios con los datos (cada dict = una fila)
        sheet_name: Nombre de la hoja de cálculo (default: "Sheet1")
        title: Título descriptivo del artefacto
        file_name: Nombre del archivo (opcional, auto-generado si no se proporciona)
        conversation_id: ID de la conversación para asociar el artefacto
        agent_id: ID del agente que genera el archivo
        metadata: Metadatos adicionales
    
    Returns:
        Dict con:
        - success: bool
        - file_path: Ruta del archivo generado
        - file_name: Nombre del archivo
        - artifact_id: ID del artefacto creado
        - rows: Número de filas
        - columns: Número de columnas
        - message: Mensaje descriptivo
    
    Examples:
        >>> data = [
        ...     {"Nombre": "Juan", "Edad": 30, "Ciudad": "Madrid"},
        ...     {"Nombre": "María", "Edad": 25, "Ciudad": "Barcelona"}
        ... ]
        >>> await generate_spreadsheet(data, title="Usuarios", sheet_name="Usuarios")
    """
    
    logger.info(
        "📊 Generating spreadsheet",
        rows=len(data),
        sheet_name=sheet_name,
        has_title=bool(title)
    )
    
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Validar datos
        if not data or not isinstance(data, list):
            return {
                "success": False,
                "error": "Se requiere una lista de datos no vacía",
                "message": "Proporciona al menos una fila de datos"
            }
        
        # Crear directorio
        WORKSPACE_PATH.mkdir(parents=True, exist_ok=True)
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if not file_name:
            safe_title = (title or "spreadsheet").replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
            file_name = f"{safe_title}_{timestamp}.xlsx"
        elif not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        
        file_path = WORKSPACE_PATH / file_name
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel limita a 31 caracteres
        
        # Obtener headers del primer elemento
        first_row = data[0]
        if not isinstance(first_row, dict):
            return {
                "success": False,
                "error": "Los datos deben ser una lista de diccionarios",
                "message": "Cada fila debe ser un objeto/diccionario con columnas"
            }
        
        headers = list(first_row.keys())
        columns_count = len(headers)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Escribir headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Escribir datos
        rows_count = 0
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Auto-detectar tipo y aplicar formato
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
            rows_count += 1
        
        # Ajustar anchos de columna
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            column_letter = get_column_letter(col_idx)
            
            # Revisar todas las celdas de la columna
            for row in ws[column_letter]:
                try:
                    cell_length = len(str(row.value)) if row.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Ajustar ancho (mínimo 10, máximo 50, con padding)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ajustar altura de la fila de headers
        ws.row_dimensions[1].height = 25
        
        # Guardar archivo
        wb.save(file_path)
        file_size = file_path.stat().st_size
        
        logger.info(
            "✅ Excel file created",
            file_path=str(file_path),
            rows=rows_count,
            columns=columns_count,
            file_size=file_size
        )
        
        # Crear artefacto
        try:
            from src.artifacts import ArtifactRepository, ArtifactCreate, ArtifactType
            
            artifact_title = title or f"Hoja de cálculo: {sheet_name}"
            artifact_description = f"Spreadsheet con {rows_count} filas y {columns_count} columnas"
            
            artifact_data = ArtifactCreate(
                type=ArtifactType.SPREADSHEET,
                title=artifact_title,
                description=artifact_description,
                file_path=str(file_path),
                file_name=file_name,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                file_size=file_size,
                conversation_id=conversation_id,
                agent_id=agent_id,
                tool_id="generate_spreadsheet",
                metadata={
                    "sheets_count": 1,
                    "rows_count": rows_count,
                    "columns_count": columns_count,
                    "file_format": "xlsx",
                    "sheet_name": sheet_name,
                    **(metadata or {})
                }
            )
            
            artifact = await ArtifactRepository.create(_user_id or "default", artifact_data)
            
            if artifact:
                logger.info(
                    "✅ Spreadsheet artifact created",
                    artifact_id=artifact.artifact_id,
                    title=artifact_title
                )
                
                return {
                    "success": True,
                    "file_path": str(file_path),
                    "file_name": file_name,
                    "artifact_id": artifact.artifact_id,
                    "rows": rows_count,
                    "columns": columns_count,
                    "message": f"✅ Excel generado exitosamente: {rows_count} filas, {columns_count} columnas. ID: {artifact.artifact_id}"
                }
            else:
                return {
                    "success": True,
                    "file_path": str(file_path),
                    "file_name": file_name,
                    "artifact_id": None,
                    "rows": rows_count,
                    "columns": columns_count,
                    "message": f"✅ Excel generado: {rows_count} filas, {columns_count} columnas (sin registro de artefacto)"
                }
                
        except Exception as e:
            logger.error(f"Error creating artifact: {e}")
            return {
                "success": True,
                "file_path": str(file_path),
                "file_name": file_name,
                "artifact_id": None,
                "rows": rows_count,
                "columns": columns_count,
                "message": f"✅ Excel generado pero error al crear artefacto: {str(e)}"
            }
        
    except ImportError:
        return {
            "success": False,
            "error": "openpyxl no está instalado",
            "message": "Instala openpyxl: pip install openpyxl"
        }
    except Exception as e:
        logger.error(f"❌ Error generating spreadsheet: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "message": f"Error al generar Excel: {str(e)}"
        }


# ============================================
# Tool Registry Definition
# ============================================

GENERATE_SPREADSHEET_TOOL = {
    "id": "generate_spreadsheet",
    "name": "generate_spreadsheet",
    "description": """Genera archivos Excel (.xlsx) a partir de datos tabulares/JSON.

Esta herramienta crea hojas de cálculo profesionales con:
- Formato automático de headers (color, negrita)
- Auto-ajuste de anchos de columna
- Formato numérico inteligente
- Primera fila congelada (para scroll)
- Bordes y estilos profesionales

Casos de uso:
- Exportar datos de análisis a Excel
- Crear reportes tabulares
- Generar datos para análisis externo
- Crear plantillas de datos

Args:
    data: Lista de objetos/diccionarios. Ej: [{"Nombre": "Juan", "Edad": 30}, ...]
    sheet_name: Nombre de la hoja (default: "Sheet1")
    title: Título descriptivo del artefacto
    file_name: Nombre de archivo (auto-generado si no se proporciona)
    metadata: Metadatos adicionales

Returns:
    Archivo Excel (.xlsx) con formato profesional, registrado como artifact tipo 'spreadsheet'.

Examples:
    >>> # Datos simples
    >>> data = [{"Producto": "Laptop", "Precio": 999.99, "Stock": 50}]
    >>> await generate_spreadsheet(data, title="Inventario")
    
    >>> # Datos complejos
    >>> data = [
    ...     {"Nombre": "Ana", "Departamento": "Ventas", "Ventas": 15000.50},
    ...     {"Nombre": "Luis", "Departamento": "Marketing", "Ventas": 12500.75}
    ... ]
    >>> await generate_spreadsheet(data, sheet_name="Ventas Q1", title="Reporte Ventas")
""",
    "parameters": {
        "type": "object",
        "properties": {
            "data": {
                "type": "array",
                "description": "Lista de objetos/diccionarios con los datos (cada objeto = una fila)",
                "items": {
                    "type": "object"
                }
            },
            "sheet_name": {
                "type": "string",
                "description": "Nombre de la hoja de cálculo",
                "default": "Sheet1"
            },
            "title": {
                "type": "string",
                "description": "Título descriptivo del artefacto"
            },
            "file_name": {
                "type": "string",
                "description": "Nombre del archivo (opcional, auto-generado si no se proporciona)"
            },
            "metadata": {
                "type": "object",
                "description": "Metadatos adicionales para el artefacto"
            }
        },
        "required": ["data"]
    },
    "handler": generate_spreadsheet
}
